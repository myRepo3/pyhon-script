import os
import google.auth
from googleapiclient import discovery
from openpyxl import Workbook

def list_instances(compute, project, zone):
    result = compute.instances().list(project=project, zone=zone).execute()
    return result.get("items", [])

def main():
    # Set up the Google Compute Engine API client
    credentials, project_id = google.auth.default()
    compute = discovery.build("compute", "v1", credentials=credentials)

    # Set the zone for your GCE instances
    zone = "us-central1-a"

    # Fetch the instance information
    instances = list_instances(compute, project_id, zone)

    # Create a new Excel workbook and add a sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "GCE Instances"

    # Write the headers to the sheet
    headers = ["Name", "ID", "Status", "Machine Type"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header

    # Write the instance information to the sheet
    for row_num, instance in enumerate(instances, 2):
        ws.cell(row=row_num, column=1).value = instance["name"]
        ws.cell(row=row_num, column=2).value = instance["id"]
        ws.cell(row=row_num, column=3).value = instance["status"]
        ws.cell(row=row_num, column=4).value = instance["machineType"].split("/")[-1]

    # Save the workbook to an Excel file
    excel_file = "gce_info.xlsx"
    wb.save(excel_file)
    print(f"Instance information saved to {excel_file}")

if __name__ == "__main__":
    main()
